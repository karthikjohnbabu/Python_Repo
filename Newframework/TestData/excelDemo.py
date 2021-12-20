import openpyxl
book= openpyxl.load_workbook("C:\\Users\\karthikb\\Desktop\\Daily\\learnwithkarthik\\Selenium-Python\\PythonDemo.xlsx")
sheet =book.active
Dict={}
cell=sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2,column=2).value="Modi"
print(sheet.cell(row=2,column=2).value)

print(sheet.max_row)
print(sheet.max_column)
#print(sheet['A3'].value)

for i in range(1,sheet.max_row+1): # Outer for loop to loop the columns
    if sheet.cell(row=i,column=1).value=="Testcase2":

        for j in range (2, sheet.max_column+1): #inner for loop to the loop rows.
           Dict[sheet.cell(row=1,column=j).value]= sheet.cell(row=i,column=j).value
print(Dict)



