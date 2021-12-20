import openpyxl


class HomePageData:
    test_HomePage_Data = [{"firstname": "karthik", "lastname": "Babu", "gender": "Male"},
                          {"firstname": "Roshini", "lastname": "John", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict={}
        book = openpyxl.load_workbook(
            "C:\\Users\\karthikb\\Desktop\\Daily\\learnwithkarthik\\Selenium-Python\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # Outer for loop to loop the columns
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # inner for loop to the loop rows.
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]

